"""Scenario export — Excel (.xlsx) and PDF, mirroring the on-screen year-by-year
tables and headline metrics. Pure functions (no Streamlit), so they're unit-testable
and reusable. The app wires them into download buttons.

All per-year figures are medians across trials — identical to what app.py shows.
"""
from __future__ import annotations

import io

import numpy as np
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Font

# Column orders mirror app.py's render_*_year_table exactly.
PROPERTY_HEADERS = ["Year", "Property value", "Loan balance", "Your equity",
                    "Net rent", "Loan interest", "Tax benefit", "Property cashflow"]
SHARES_HEADERS = ["Year", "Cash injected", "Dividends reinvested", "Capital growth",
                  "Dividend tax", "Share value"]


def _median_path(result, key, deflate, yearly_deflator):
    """Median across trials per year (matches app.py:_median_path)."""
    arr = result[key]
    if deflate and yearly_deflator is not None:
        arr = arr / yearly_deflator
    return np.median(arr, axis=0)


def _property_rows(result, horizon, deflate, yearly_deflator):
    """Return (headers, rows) with numeric cells. Tax benefit is shown as a
    benefit (negated tax, matching the on-screen `-tax`)."""
    def med(key):
        return _median_path(result, key, deflate, yearly_deflator)

    value, balance = med("property_value_path"), med("property_loan_balance_path")
    rent, interest = med("property_rent_path"), med("property_interest_path")
    tax, pcash = med("property_tax_path"), med("property_cashflow_path")
    rows = []
    for i in range(horizon):
        rows.append([
            i + 1,
            float(value[i]), float(balance[i]), float(value[i] - balance[i]),
            float(rent[i]), float(interest[i]), float(-tax[i]), float(pcash[i]),
        ])
    return PROPERTY_HEADERS, rows


def _shares_rows(result, horizon, mix_pct, deflate, yearly_deflator):
    """Return (headers, rows). Adds a 'Mix wealth' column when mix_pct < 100."""
    def med(key):
        return _median_path(result, key, deflate, yearly_deflator)

    inj, div = med("shares_contribution_path"), med("shares_dividend_path")
    grow, dtax = med("shares_capital_growth_path"), med("shares_dividend_tax_path")
    sval = med("shares_wealth_path")
    show_mix = mix_pct < 100
    mixw = med("mixed_wealth_path") if show_mix else None

    headers = list(SHARES_HEADERS) + (["Mix wealth"] if show_mix else [])
    rows = []
    for i in range(horizon):
        row = [i + 1, float(inj[i]), float(div[i]), float(grow[i]),
               float(dtax[i]), float(sval[i])]
        if show_mix:
            row.append(float(mixw[i]))
        rows.append(row)
    return headers, rows


def _summary_pairs(result, meta):
    """Label/value pairs for the summary block — inputs + headline metrics."""
    return [
        ("Purchase price", f"${meta['purchase_price']:,.0f}"),
        ("Deposit", f"{meta['deposit_pct']}%"),
        ("Gross rent yield", f"{meta['gross_yield_pct']:.1f}%"),
        ("Loan rate", f"{meta['loan_rate_pct']:.1f}%"),
        ("Horizon", f"{meta['horizon']} years"),
        ("State", str(meta.get("state", ""))),
        ("Taxable income", f"${meta['income']:,.0f}"),
        ("Marginal tax rate", f"{meta['mtr']*100:.0f}%"),
        ("Property / shares mix", f"{meta['property_share_mix_pct']}% / "
                                  f"{100 - meta['property_share_mix_pct']}%"),
        ("Share portfolio", str(meta["portfolio_profile"])),
        ("Property regime", str(meta["property_regime"])),
        ("Max annual top-up planned", f"${meta['max_top_up']:,.0f}"),
        ("—", "—"),
        ("Typical property wealth", f"${result['median_property_wealth']:,.0f}"),
        ("Typical shares wealth", f"${result['median_shares_wealth']:,.0f}"),
        ("Worst-year cash needed", f"${meta['worst_year_cash']:,.0f}"),
        ("Chance you never run out of cash", f"{result['p_solvent']*100:.0f}%"),
    ]


# ===========================================================================
# Excel
# ===========================================================================
def build_excel(result, meta, horizon, mix_pct, deflate, yearly_deflator) -> bytes:
    wb = Workbook()
    bold = Font(bold=True)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Property vs Shares — scenario"
    ws["A1"].font = Font(bold=True, size=14)
    r = 3
    for label, val in _summary_pairs(result, meta):
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # Year-by-year sheets (numeric cells so the user can compute)
    for title, (headers, rows) in [
        ("Property year-by-year", _property_rows(result, horizon, deflate, yearly_deflator)),
        ("Shares year-by-year", _shares_rows(result, horizon, mix_pct, deflate, yearly_deflator)),
    ]:
        sheet = wb.create_sheet(title)
        for c, h in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=c, value=h)
            cell.font = bold
        for ri, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                cell = sheet.cell(row=ri, column=c, value=val)
                if c > 1:
                    cell.number_format = "#,##0"
        for c in range(1, len(headers) + 1):
            sheet.column_dimensions[sheet.cell(row=1, column=c).column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# PDF
# ===========================================================================
def _money(x: float) -> str:
    return f"-${abs(x):,.0f}" if x < 0 else f"${x:,.0f}"


def _pdf_table(pdf: FPDF, headers, rows):
    n = len(headers)
    usable = pdf.w - pdf.l_margin - pdf.r_margin
    first = usable * 0.08
    rest = (usable - first) / (n - 1)
    widths = [first] + [rest] * (n - 1)

    pdf.set_font("Helvetica", "B", 7.5)
    pdf.set_fill_color(243, 244, 246)
    for w, h in zip(widths, headers):
        pdf.cell(w, 7, h, border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7.5)
    for row in rows:
        for c, (w, val) in enumerate(zip(widths, row)):
            text = str(val) if c == 0 else _money(val)
            pdf.cell(w, 6, text, border=1, align="C" if c == 0 else "R")
        pdf.ln()


def build_pdf(result, meta, horizon, mix_pct, deflate, yearly_deflator) -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    nl = dict(new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Property vs Shares - scenario", **nl)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(110, 110, 110)
    pdf.cell(0, 6, "What-if explorer for personal use. Not financial advice. "
                   "Figures are medians across 5,000 simulated futures.", **nl)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    # Summary block — two columns of label/value pairs (drop the divider row,
    # which uses an em-dash the PDF core font can't encode).
    pairs = [p for p in _summary_pairs(result, meta) if p[0] != "—"]
    pdf.set_font("Helvetica", "", 9)
    col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / 2
    half = (len(pairs) + 1) // 2
    left, right = pairs[:half], pairs[half:]
    for i in range(half):
        for col, pair in ((0, left[i] if i < len(left) else None),
                          (1, right[i] if i < len(right) else None)):
            if pair is None:
                pdf.cell(col_w, 6, "")
                continue
            label, val = pair
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(col_w * 0.55, 6, f"{label}")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(col_w * 0.45, 6, f"{val}")
        pdf.ln()

    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Property - year by year", **nl)
    _pdf_table(pdf, *_property_rows(result, horizon, deflate, yearly_deflator))

    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Shares - year by year", **nl)
    _pdf_table(pdf, *_shares_rows(result, horizon, mix_pct, deflate, yearly_deflator))

    out = pdf.output()
    return bytes(out)
