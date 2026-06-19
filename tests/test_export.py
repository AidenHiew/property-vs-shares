import io

import numpy as np
import pytest
import openpyxl

from ui.export import (
    _median_path, _property_rows, _shares_rows,
    build_excel, build_pdf,
)


def _fake_result(horizon=4, trials=10):
    rng = np.random.default_rng(0)

    def path(scale):
        return rng.uniform(0.5, 1.5, (trials, horizon)) * scale

    return {
        "property_value_path": path(800_000),
        "property_loan_balance_path": path(500_000),
        "property_rent_path": path(28_000),
        "property_interest_path": path(31_000),
        "property_tax_path": path(4_000),
        "property_cashflow_path": path(-3_000),
        "shares_contribution_path": path(3_000),
        "shares_dividend_path": path(2_000),
        "shares_capital_growth_path": path(20_000),
        "shares_dividend_tax_path": path(-600),
        "shares_wealth_path": path(300_000),
        "mixed_wealth_path": path(350_000),
        "median_property_wealth": 1_800_000.0,
        "median_shares_wealth": 1_200_000.0,
        "p_solvent": 0.96,
    }


def _meta():
    return {
        "purchase_price": 700_000, "deposit_pct": 20, "gross_yield_pct": 4.0,
        "income": 150_000, "mtr": 0.37, "loan_rate_pct": 6.0, "horizon": 4,
        "max_top_up": 20_000, "property_share_mix_pct": 60,
        "portfolio_profile": "blended", "property_regime": "current",
        "state": "SA", "worst_year_cash": 18_000.0,
    }


# ---------------------------------------------------------------------------
# row helpers
# ---------------------------------------------------------------------------
def test_median_path_reduces_trials_axis():
    r = _fake_result(horizon=4, trials=10)
    med = _median_path(r, "property_value_path", deflate=False, yearly_deflator=None)
    assert med.shape == (4,)
    assert np.allclose(med, np.median(r["property_value_path"], axis=0))


def test_property_rows_equity_is_value_minus_balance():
    r = _fake_result(horizon=4)
    headers, rows = _property_rows(r, horizon=4, deflate=False, yearly_deflator=None)
    assert headers[0] == "Year"
    assert "Your equity" in headers
    assert len(rows) == 4
    # row = [year, value, balance, equity, rent, interest, tax_benefit, cashflow]
    for row in rows:
        value, balance, equity = row[1], row[2], row[3]
        assert equity == pytest.approx(value - balance)


def test_shares_rows_include_mix_when_mix_below_100():
    r = _fake_result(horizon=4)
    headers, rows = _shares_rows(r, horizon=4, mix_pct=60, deflate=False, yearly_deflator=None)
    assert "Mix wealth" in headers
    assert len(rows[0]) == len(headers)


def test_shares_rows_drop_mix_when_pure():
    r = _fake_result(horizon=4)
    headers, rows = _shares_rows(r, horizon=4, mix_pct=100, deflate=False, yearly_deflator=None)
    assert "Mix wealth" not in headers


# ---------------------------------------------------------------------------
# excel
# ---------------------------------------------------------------------------
def test_build_excel_is_valid_xlsx_with_expected_sheets():
    r, m = _fake_result(), _meta()
    data = build_excel(r, m, horizon=4, mix_pct=60, deflate=False, yearly_deflator=None)
    assert data[:2] == b"PK"  # xlsx is a zip
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Property year-by-year" in wb.sheetnames
    assert "Shares year-by-year" in wb.sheetnames


def test_build_excel_property_sheet_has_header_plus_year_rows():
    r, m = _fake_result(horizon=4), _meta()
    data = build_excel(r, m, horizon=4, mix_pct=60, deflate=False, yearly_deflator=None)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Property year-by-year"]
    assert ws.max_row == 5  # 1 header + 4 years
    assert ws.cell(row=1, column=1).value == "Year"
    # numeric, not a pre-formatted string (so the user can compute on it)
    assert isinstance(ws.cell(row=2, column=2).value, (int, float))


# ---------------------------------------------------------------------------
# pdf
# ---------------------------------------------------------------------------
def test_build_pdf_is_valid_pdf_bytes():
    r, m = _fake_result(), _meta()
    data = build_pdf(r, m, horizon=4, mix_pct=60, deflate=False, yearly_deflator=None)
    assert data[:4] == b"%PDF"
    assert len(data) > 1000
